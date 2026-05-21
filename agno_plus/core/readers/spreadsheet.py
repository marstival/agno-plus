"""SpreadsheetReader — layout-aware Excel/CSV ingestion.

3-layer pipeline ported from the reference personal-agentic-aide codebase:
  Layer A (grid reader)     → sparse SheetGrid, merged-cell expansion
  Layer B (block detector)  → BFS connected-component classification
  Layer C (record extractor)→ TableRow / KVItem / NoteItem records

Outputs one Document per detected block (not one per sheet). Agno's native
ExcelReader uses read_only=True and silently drops merged cell content; this
reader uses read_only=False and propagates master cell values.
"""

from __future__ import annotations

import csv
import io
import logging
import uuid
from collections import deque
from pathlib import Path
from statistics import stdev
from typing import Any

from agno_plus.core.models import Document
from agno_plus.core.readers._spreadsheet_models import (
    Block,
    BlockType,
    CellValue,
    KVItem,
    NoteItem,
    ParseResult,
    SheetGrid,
    TableRow,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Detection thresholds (constants ported from reference)
# ---------------------------------------------------------------------------

TABLE_MIN_DIMENSION = 3
KV_PAIR_WIDTH = 2
KV_PAIR_MIN_HEIGHT = 2
KV_PAIR_MIN_LEFT_FILL = 0.8
KV_PAIR_HIGH_LEFT_FILL = 0.9
KV_PAIR_CONFIDENCE_PENALTY = 0.1
NOTE_BLOCK_DEFAULT_CONFIDENCE = 0.8
TABLE_TOP_FILL_MIN = 0.6
TABLE_TOP_FILL_PENALTY_BELOW = 0.7
TABLE_CONFIDENCE_PENALTY = 0.1
TABLE_FILL_STD_REJECTION = 0.3
TABLE_FILL_STD_PENALTY = 0.2

SUPPORTED_EXTENSIONS = frozenset({".xlsx", ".xls", ".xlsm", ".csv", ".tsv"})


# ---------------------------------------------------------------------------
# SpreadsheetReader
# ---------------------------------------------------------------------------


class SpreadsheetReader:
    """Deterministic layout-aware reader for Excel and CSV files.

    Returns one Document per detected block. Each Document carries structured
    Markdown content (tables, key-value pairs, notes) and block-level metadata.
    """

    def read(self, source: bytes | str, filename: str = "upload.xlsx", **kwargs: Any) -> list[Document]:
        if isinstance(source, str):
            source = source.encode()
        result = self._parse(source, filename)
        return self._to_documents(result)

    def is_supported(self, filename: str) -> bool:
        return Path(filename).suffix.lower() in SUPPORTED_EXTENSIONS

    def extract_tables(
        self, source: bytes | str, filename: str = "upload.xlsx"
    ) -> list[dict]:
        """Return raw table data for structured domain ingestion.

        Each entry: {"headers": [...], "rows": [{col: val, ...}, ...]}.
        Only TABLE blocks are returned (KV_PAIR and NOTE blocks are skipped).
        """
        if isinstance(source, str):
            source = source.encode()
        result = self._parse(source, filename)

        tables: dict[int, dict] = {}
        for block in result.blocks:
            if block.block_type == BlockType.TABLE:
                tables[result.blocks.index(block)] = {
                    "headers": list(block.headers),
                    "rows": [],
                }

        for row in result.table_rows:
            if row.block_index in tables:
                tables[row.block_index]["rows"].append(dict(row.fields))

        return list(tables.values())

    # ------------------------------------------------------------------
    # Internal: orchestrate A → B → C → Document
    # ------------------------------------------------------------------

    def _parse(self, file_bytes: bytes, filename: str) -> ParseResult:
        grids = _layer_a_read(file_bytes, filename)

        all_blocks: list[Block] = []
        all_table_rows: list[TableRow] = []
        all_kv_items: list[KVItem] = []
        all_notes: list[NoteItem] = []

        for grid in grids:
            block_offset = len(all_blocks)
            blocks = _layer_b_detect(grid)
            table_rows, kv_items, notes = _layer_c_extract(blocks, grid)
            # Offset block_index references so they remain globally unique
            for row in table_rows:
                row.block_index += block_offset
            for item in kv_items:
                item.block_index += block_offset
            for note in notes:
                note.block_index += block_offset
            all_blocks.extend(blocks)
            all_table_rows.extend(table_rows)
            all_kv_items.extend(kv_items)
            all_notes.extend(notes)

        logger.debug(
            "Parsed %s: %d sheets, %d blocks, %d rows, %d kvs, %d notes",
            filename, len(grids), len(all_blocks),
            len(all_table_rows), len(all_kv_items), len(all_notes),
        )
        return ParseResult(
            filename=filename,
            sheet_grids=grids,
            blocks=all_blocks,
            table_rows=all_table_rows,
            kv_items=all_kv_items,
            notes=all_notes,
        )

    def _to_documents(self, result: ParseResult) -> list[Document]:
        docs: list[Document] = []
        for block_idx, block in enumerate(result.blocks):
            content = _render_block(
                block,
                block_idx,
                result.table_rows,
                result.kv_items,
                result.notes,
            )
            if not content:
                continue
            docs.append(
                Document(
                    id=f"spreadsheet-{uuid.uuid4().hex[:8]}",
                    content=content,
                    source_type="spreadsheet",
                    source_name=result.filename,
                    metadata={
                        "block_type": block.block_type.value,
                        "sheet_name": block.sheet_name,
                        "top_row": block.top_row,
                        "bottom_row": block.bottom_row,
                        "left_col": block.left_col,
                        "right_col": block.right_col,
                        "confidence": block.confidence,
                        "block_index": block_idx,
                    },
                )
            )
        return docs


# ---------------------------------------------------------------------------
# Layer A: grid reader
# ---------------------------------------------------------------------------


def _layer_a_read(file_bytes: bytes, filename: str) -> list[SheetGrid]:
    ext = Path(filename).suffix.lower()
    if ext in {".xlsx", ".xls", ".xlsm"}:
        return _read_excel(file_bytes)
    if ext in {".csv", ".tsv"}:
        return _read_csv(file_bytes, ext)
    raise ValueError(f"Unsupported extension: {ext}")


def _read_excel(file_bytes: bytes) -> list[SheetGrid]:
    import openpyxl

    buf = io.BytesIO(file_bytes)
    wb = openpyxl.load_workbook(buf, data_only=True, read_only=False)
    grids: list[SheetGrid] = []
    for ws in wb.worksheets:
        cells = _extract_cells_with_merges(ws)
        grids.append(
            SheetGrid(
                sheet_name=ws.title,
                cells=cells,
                row_count=ws.max_row or 0,
                col_count=ws.max_column or 0,
            )
        )
    wb.close()
    return grids


def _extract_cells_with_merges(ws: Any) -> list[CellValue]:
    merged_map: dict[tuple[int, int], str] = {}
    merged_coords: set[tuple[int, int]] = set()

    for merged_range in ws.merged_cells.ranges:
        master_row = merged_range.min_row
        master_col = merged_range.min_col
        master_value = _cell_to_str(ws.cell(row=master_row, column=master_col).value)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if row == master_row and col == master_col:
                    continue
                merged_map[(row - 1, col - 1)] = master_value
                merged_coords.add((row - 1, col - 1))

    cells: list[CellValue] = []
    for row in ws.iter_rows():
        for cell in row:
            r, c = cell.row - 1, cell.column - 1
            if (r, c) in merged_coords:
                value = merged_map[(r, c)]
                if value:
                    cells.append(CellValue(row=r, col=c, value=value, merged=True))
            else:
                value = _cell_to_str(cell.value)
                if value:
                    cells.append(CellValue(row=r, col=c, value=value))
    return cells


def _cell_to_str(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _read_csv(file_bytes: bytes, ext: str) -> list[SheetGrid]:
    text = _decode_bytes(file_bytes)
    delimiter = _sniff_delimiter(text, ext)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    cells: list[CellValue] = []
    max_row = max_col = 0
    for row_idx, row in enumerate(reader):
        for col_idx, value in enumerate(row):
            stripped = value.strip()
            if stripped:
                cells.append(CellValue(row=row_idx, col=col_idx, value=stripped))
                max_col = max(max_col, col_idx)
        if row:
            max_row = row_idx
    return [SheetGrid(sheet_name="Sheet1", cells=cells, row_count=max_row + 1, col_count=max_col + 1)]


def _decode_bytes(file_bytes: bytes) -> str:
    try:
        return file_bytes.decode("utf-8")
    except UnicodeDecodeError:
        return file_bytes.decode("latin-1")


def _sniff_delimiter(text: str, ext: str) -> str:
    if ext == ".tsv":
        return "\t"
    try:
        return csv.Sniffer().sniff(text[:8192], delimiters=",;\t|").delimiter
    except csv.Error:
        return ","


# ---------------------------------------------------------------------------
# Layer B: block detector
# ---------------------------------------------------------------------------


def _layer_b_detect(grid: SheetGrid, gap: int = 1) -> list[Block]:
    if not grid.cells:
        return []
    occupied = {(c.row, c.col) for c in grid.cells}
    cell_map = {(c.row, c.col): c for c in grid.cells}
    components = _find_connected_components(occupied, gap)
    blocks = [_classify_component(coords, cell_map, grid.sheet_name) for coords in components]
    blocks.sort(key=lambda b: (b.top_row, b.left_col))
    return blocks


def _find_connected_components(
    occupied: set[tuple[int, int]], gap: int
) -> list[list[tuple[int, int]]]:
    visited: set[tuple[int, int]] = set()
    components: list[list[tuple[int, int]]] = []
    reach = gap + 1

    for start in occupied:
        if start in visited:
            continue
        component: list[tuple[int, int]] = []
        queue: deque[tuple[int, int]] = deque([start])
        visited.add(start)
        while queue:
            cr, cc = queue.popleft()
            component.append((cr, cc))
            for dr in range(-reach, reach + 1):
                for dc in range(-reach, reach + 1):
                    if dr == 0 and dc == 0:
                        continue
                    nr, nc = cr + dr, cc + dc
                    if (nr, nc) in occupied and (nr, nc) not in visited:
                        visited.add((nr, nc))
                        queue.append((nr, nc))
        components.append(component)
    return components


def _classify_component(
    coords: list[tuple[int, int]],
    cell_map: dict[tuple[int, int], CellValue],
    sheet_name: str,
) -> Block:
    top_row = min(r for r, _ in coords)
    left_col = min(c for _, c in coords)
    bottom_row = max(r for r, _ in coords)
    right_col = max(c for _, c in coords)
    width = right_col - left_col + 1
    height = bottom_row - top_row + 1
    cells = [cell_map[(r, c)] for r, c in coords if (r, c) in cell_map]
    coord_set = set(coords)

    if width >= TABLE_MIN_DIMENSION and height >= TABLE_MIN_DIMENSION:
        block_type, has_header, headers, confidence = _try_table(
            top_row, left_col, bottom_row, right_col, width, height, coord_set
        )
        if block_type == BlockType.TABLE:
            return Block(
                block_type=block_type, sheet_name=sheet_name,
                top_row=top_row, left_col=left_col,
                bottom_row=bottom_row, right_col=right_col,
                cells=cells, has_header=has_header,
                headers=headers, confidence=confidence,
            )

    if width == KV_PAIR_WIDTH and height >= KV_PAIR_MIN_HEIGHT:
        left_fill = sum(1 for r, c in coords if c == left_col) / height
        if left_fill >= KV_PAIR_MIN_LEFT_FILL:
            conf = 1.0 if left_fill >= KV_PAIR_HIGH_LEFT_FILL else 1.0 - KV_PAIR_CONFIDENCE_PENALTY
            return Block(
                block_type=BlockType.KV_PAIR, sheet_name=sheet_name,
                top_row=top_row, left_col=left_col,
                bottom_row=bottom_row, right_col=right_col,
                cells=cells, confidence=conf,
            )

    return Block(
        block_type=BlockType.NOTE, sheet_name=sheet_name,
        top_row=top_row, left_col=left_col,
        bottom_row=bottom_row, right_col=right_col,
        cells=cells, confidence=NOTE_BLOCK_DEFAULT_CONFIDENCE,
    )


def _try_table(
    top_row: int, left_col: int, bottom_row: int, right_col: int,
    width: int, height: int,
    coord_set: set[tuple[int, int]],
) -> tuple[BlockType, bool, list[str], float]:
    confidence = 1.0
    top_fill = sum(1 for c in range(left_col, right_col + 1) if (top_row, c) in coord_set) / width
    if top_fill < TABLE_TOP_FILL_MIN:
        return BlockType.NOTE, False, [], NOTE_BLOCK_DEFAULT_CONFIDENCE
    if top_fill < TABLE_TOP_FILL_PENALTY_BELOW:
        confidence -= TABLE_CONFIDENCE_PENALTY

    data_fills = [
        sum(1 for c in range(left_col, right_col + 1) if (r, c) in coord_set) / width
        for r in range(top_row + 1, bottom_row + 1)
        if any((r, c) in coord_set for c in range(left_col, right_col + 1))
    ]
    if not data_fills:
        return BlockType.NOTE, False, [], NOTE_BLOCK_DEFAULT_CONFIDENCE
    if len(data_fills) >= 2:
        fill_std = stdev(data_fills)
        if fill_std >= TABLE_FILL_STD_REJECTION:
            return BlockType.NOTE, False, [], NOTE_BLOCK_DEFAULT_CONFIDENCE
        if fill_std > TABLE_FILL_STD_PENALTY:
            confidence -= TABLE_CONFIDENCE_PENALTY

    return BlockType.TABLE, True, [f"col_{i}" for i in range(width)], confidence


# ---------------------------------------------------------------------------
# Layer C: record extractor
# ---------------------------------------------------------------------------


def _layer_c_extract(
    blocks: list[Block], grid: SheetGrid
) -> tuple[list[TableRow], list[KVItem], list[NoteItem]]:
    cell_map = {(c.row, c.col): c.value for c in grid.cells}
    table_rows: list[TableRow] = []
    kv_items: list[KVItem] = []
    notes: list[NoteItem] = []
    for idx, block in enumerate(blocks):
        if block.block_type == BlockType.TABLE:
            table_rows.extend(_extract_table_rows(block, idx, cell_map))
        elif block.block_type == BlockType.KV_PAIR:
            kv_items.extend(_extract_kv_items(block, idx, cell_map))
        elif block.block_type == BlockType.NOTE:
            note = _extract_note(block, idx, cell_map)
            if note:
                notes.append(note)
    return table_rows, kv_items, notes


def _extract_table_rows(
    block: Block, block_idx: int, cell_map: dict[tuple[int, int], str]
) -> list[TableRow]:
    width = block.right_col - block.left_col + 1
    if block.has_header:
        headers = [
            cell_map.get((block.top_row, c), "") or f"col_{c - block.left_col}"
            for c in range(block.left_col, block.right_col + 1)
        ]
        data_start = block.top_row + 1
    else:
        headers = block.headers or [f"col_{i}" for i in range(width)]
        data_start = block.top_row

    # Deduplicate headers
    seen: dict[str, int] = {}
    unique: list[str] = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            unique.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            unique.append(h)

    rows: list[TableRow] = []
    row_idx = 0
    for r in range(data_start, block.bottom_row + 1):
        fields = {
            unique[i] if i < len(unique) else f"col_{i}": cell_map.get((r, c), "")
            for i, c in enumerate(range(block.left_col, block.right_col + 1))
        }
        if any(fields.values()):
            rows.append(TableRow(
                sheet_name=block.sheet_name,
                block_index=block_idx,
                row_index=row_idx,
                fields=fields,
            ))
            row_idx += 1
    return rows


def _extract_kv_items(
    block: Block, block_idx: int, cell_map: dict[tuple[int, int], str]
) -> list[KVItem]:
    items: list[KVItem] = []
    for r in range(block.top_row, block.bottom_row + 1):
        key = cell_map.get((r, block.left_col), "").strip()
        value = cell_map.get((r, block.left_col + 1), "").strip()
        if key:
            items.append(KVItem(sheet_name=block.sheet_name, block_index=block_idx, key=key, value=value))
    return items


def _extract_note(
    block: Block, block_idx: int, cell_map: dict[tuple[int, int], str]
) -> NoteItem | None:
    texts = [
        cell_map.get((r, c), "").strip()
        for r in range(block.top_row, block.bottom_row + 1)
        for c in range(block.left_col, block.right_col + 1)
        if cell_map.get((r, c), "").strip()
    ]
    if not texts:
        return None
    return NoteItem(sheet_name=block.sheet_name, block_index=block_idx, text=" ".join(texts))


# ---------------------------------------------------------------------------
# Text renderer: block → Markdown (used by _to_documents)
# ---------------------------------------------------------------------------


def _render_block(
    block: Block,
    block_idx: int,
    table_rows: list[TableRow],
    kv_items: list[KVItem],
    notes: list[NoteItem],
) -> str:
    col_range = f"{_col_letter(block.left_col)}-{_col_letter(block.right_col)}"
    row_range = f"{block.top_row + 1}-{block.bottom_row + 1}"

    if block.block_type == BlockType.TABLE:
        rows = [r for r in table_rows if r.block_index == block_idx]
        return _render_table(f"### Table (rows {row_range}, cols {col_range})", rows)
    if block.block_type == BlockType.KV_PAIR:
        items = [k for k in kv_items if k.block_index == block_idx]
        return _render_kv(f"### Key-Value (rows {row_range}, cols {col_range})", items)
    if block.block_type == BlockType.NOTE:
        block_notes = [n for n in notes if n.block_index == block_idx]
        return _render_notes(f"### Note (rows {row_range}, cols {col_range})", block_notes)
    return ""


def _render_table(heading: str, rows: list[TableRow]) -> str:
    if not rows:
        return ""
    headers = list(rows[0].fields.keys())
    lines = [
        heading,
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
        *("| " + " | ".join(row.fields.get(h, "") for h in headers) + " |" for row in rows),
    ]
    return "\n".join(lines)


def _render_kv(heading: str, items: list[KVItem]) -> str:
    if not items:
        return ""
    return "\n".join([heading] + [f"- {i.key}: {i.value}" for i in items])


def _render_notes(heading: str, notes: list[NoteItem]) -> str:
    if not notes:
        return ""
    return "\n".join([heading] + [n.text for n in notes])


def _col_letter(col_index: int) -> str:
    """Convert 0-based column index to Excel letter (0→A, 25→Z, 26→AA)."""
    result = ""
    idx = col_index
    while True:
        result = chr(65 + idx % 26) + result
        idx = idx // 26 - 1
        if idx < 0:
            break
    return result
